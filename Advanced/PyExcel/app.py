import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]
cell = sheet["a1"]
column = sheet["a1"]
cell = sheet["a:c"]
sheet["a1:c3"]
sheet[1:3]

sheet.append([1, 2, 3])
wb.save("transcations2.xlsx")

# for row in range(1, sheet.max_row + 1):
#   for column in range(1, sheet.max_column + 1):
#     cell = sheet.cell(row, column)
#     print(cell.value)

# print(sheet.max_row)
# print(sheet.max_column)
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)

# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)